# -*- coding: utf-8 -*-
"""
kakao_engine.py - 카카오톡 대화방 목록 마우스 포커스 락 & 정밀 전송 v39.0
"""

import os
import time
import shutil
import tempfile
import random
import datetime
import openpyxl
import pyautogui
import pyperclip

# pyautogui 안전 딜레이
pyautogui.PAUSE = 0.05
pyautogui.FAILSAFE = False

def set_clipboard_text(text):
    """클립보드에 텍스트 복사"""
    for _ in range(5):
        try:
            pyperclip.copy(text)
            time.sleep(0.05)
            if pyperclip.paste() == text:
                return True
        except Exception:
            time.sleep(0.05)
    return True

def fix_excel_path(excel_path):
    if not os.path.exists(excel_path):
        return excel_path
    tmp_path = excel_path.replace(".xlsx", ".tmp_read.xlsx")
    try:
        shutil.copy2(excel_path, tmp_path)
        return tmp_path
    except Exception:
        return excel_path

def create_sample_excel(filepath):
    os.makedirs(os.path.dirname(os.path.abspath(filepath)), exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "고객명단"
    headers = ["고객성함", "카톡이름/전화번호", "전송할메시지", "전송상태", "전송시간", "결과메시지"]
    ws.append(headers)
    samples = [
        ["홍길동", "홍길동", "고객님, 즐거운 하루 되세요! 😊", "", "", ""],
        ["이순신", "이순신", "고객님, 이번 주 특별 혜택 안내 드립니다.", "", "", ""],
        ["강감찬", "01012345678", "안녕하세요 고객님, 신한라이프입니다.", "", "", ""]
    ]
    for row in samples:
        ws.append(row)
    wb.save(filepath)

def load_customer_list(filepath):
    real_path = fix_excel_path(filepath)
    if not os.path.exists(real_path):
        return []
    try:
        wb = openpyxl.load_workbook(real_path, data_only=True)
        ws = wb.active
        customers = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not any(row):
                continue
            name = str(row[0]).strip() if row[0] is not None else ""
            kakao_id = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
            msg = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
            status = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ""
            if name or kakao_id:
                customers.append({
                    "row_idx": row_idx,
                    "name": name,
                    "kakao_name": kakao_id if kakao_id else name,
                    "msg": msg,
                    "status": status
                })
        wb.close()
        return customers
    except Exception as e:
        print("Excel read error:", e)
        return []

def update_customer_status(filepath, row_idx, status, timestamp, result_msg):
    real_path = fix_excel_path(filepath)
    orig_path = filepath
    target = orig_path if os.path.exists(orig_path) else real_path
    try:
        wb = openpyxl.load_workbook(target)
        ws = wb.active
        ws.cell(row=row_idx, column=4, value=status)
        ws.cell(row=row_idx, column=5, value=timestamp)
        ws.cell(row=row_idx, column=6, value=result_msg)
        wb.save(target)
        wb.close()
        return True
    except Exception as e:
        print("Excel save error:", e)
        return False

def send_message_to_current_open_room(msg_text):
    pyautogui.press('enter')
    time.sleep(0.7)

    set_clipboard_text(msg_text)
    time.sleep(0.1)
    pyautogui.hotkey('ctrl', 'v')
    time.sleep(0.25)

    pyautogui.press('enter')
    time.sleep(0.35)

    pyautogui.press('esc')
    time.sleep(0.25)
    return True, "전송 완료"

def send_kakao_message_to_folder_index(index, msg_template, is_semi_auto=False):
    if index > 1:
        pyautogui.press('down')
        time.sleep(0.2)

    ok, res = send_message_to_current_open_room(msg_template)
    if not ok:
        return False, res

    return True, f"[{index}번째 대화방] 메시지 성공적 전송 완료"

def send_kakao_message_to_favorites_index(index, msg_template, is_semi_auto=False):
    return send_kakao_message_to_folder_index(index, msg_template, is_semi_auto)

def send_kakao_message_to_user(user_id_or_name, msg_text, is_semi_auto=False):
    target_name = user_id_or_name.strip()
    if not target_name:
        return False, "검색할 고객 이름이 없습니다."

    pyautogui.press('esc')
    time.sleep(0.1)

    pyautogui.hotkey('ctrl', 'f')
    time.sleep(0.25)

    pyautogui.hotkey('ctrl', 'a')
    time.sleep(0.05)
    pyautogui.press('backspace')
    time.sleep(0.05)

    set_clipboard_text(target_name)
    pyautogui.hotkey('ctrl', 'v')
    time.sleep(0.5)

    pyautogui.press('enter')
    time.sleep(0.7)

    set_clipboard_text(msg_text)
    time.sleep(0.1)
    pyautogui.hotkey('ctrl', 'v')
    time.sleep(0.25)

    if is_semi_auto:
        return True, "메시지 입력 완료 (반자동)"

    pyautogui.press('enter')
    time.sleep(0.35)

    pyautogui.press('esc')
    time.sleep(0.2)
    return True, "전송 완료"

def send_kakao_message_to_self(msg_text, is_semi_auto=False):
    return send_kakao_message_to_user("강동우", msg_text, is_semi_auto)