# -*- coding: utf-8 -*-
"""
====================================================================
카카오톡 고객 안전 자동관리 프로그램 v2.0 (노트북/데스크톱 완벽 대응)
====================================================================
[프로그램 특징]
1. 화면 배율(100%, 125%, 150%) 및 해상도에 영향받지 않는 Win32 API 창 제어 방식
2. 클립보드(pyperclip) 기반 한글/이모티콘 깨짐 없는 안전 전송
3. 고객 명단 엑셀(고객명단_양식.xlsx) 실시간 읽기 및 발송 결과 자동 기록
4. 세련된 Dark Mode GUI 대시보드 및 실시간 발송 로그 제공
5. 차단 방지를 위한 자연스러운 인간형 랜덤 지연(Delay) 시스템
====================================================================
"""

import os
import sys
import time
import json
import random
import threading
import datetime
import ctypes
import tkinter as tk
from tkinter import ttk, messagebox, filedialog

# Windows DPI 배율 인식 설정 (노트북 125%/150% 화면에서도 글씨와 UI가 선명하고 정확하게 동작하도록 설정)
try:
    ctypes.windll.shcore.SetProcessDpiAwareness(1)
except Exception:
    try:
        ctypes.windll.user32.SetProcessDPIAware()
    except Exception:
        pass

# 필수 외부 라이브러리 불러오기
try:
    import win32gui
    import win32con
    import win32api
    import win32clipboard
    import pyperclip
    import pyautogui
    import openpyxl
except ImportError as e:
    print(f"[경고] 필수 라이브러리가 설치되지 않았습니다: {e}")
    print("01_최초설치하기.bat 파일을 실행하여 패키지를 먼저 설치해 주세요.")

# 전역 설정 파일 기본 경로
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_FILE = os.path.join(BASE_DIR, "config.json")
EXCEL_DEFAULT_FILE = os.path.join(BASE_DIR, "고객명단_양식.xlsx")


class KakaoAutomationEngine:
    """카카오톡 창 제어 및 메시지 발송을 담당하는 핵심 엔진 클래스"""

    def __init__(self, log_callback=None):
        self.log_callback = log_callback
        self.is_running = False
        self.is_paused = False

    def log(self, message):
        """로그 출력 함수"""
        now = datetime.datetime.now().strftime("[%H:%M:%S]")
        formatted_msg = f"{now} {message}"
        print(formatted_msg)
        if self.log_callback:
            self.log_callback(formatted_msg)

    def find_kakao_main_window(self):
        """카카오톡 메인 창 핸들을 찾습니다."""
        hwnd = win32gui.FindWindow("EVA_Window_Client", "카카오톡")
        if not hwnd:
            # 창 이름이 다를 수 있으므로 클래스명으로 다시 탐색
            def enum_cb(h, extra):
                if win32gui.IsWindowVisible(h):
                    cls_name = win32gui.GetClassName(h)
                    title = win32gui.GetWindowText(h)
                    if cls_name == "EVA_Window_Client" and "카카오톡" in title:
                        extra.append(h)
            found = []
            win32gui.EnumWindows(enum_cb, found)
            if found:
                hwnd = found[0]
        return hwnd

    def bring_to_front(self, hwnd):
        """지정한 창을 최상단으로 가져오고 활성화합니다."""
        try:
            if win32gui.IsIconic(hwnd):
                win32gui.ShowWindow(hwnd, win32con.SW_RESTORE)
            else:
                win32gui.ShowWindow(hwnd, win32con.SW_SHOW)
            win32gui.SetForegroundWindow(hwnd)
            time.sleep(0.3)
            return True
        except Exception as e:
            self.log(f"⚠️ 창 활성화 중 알림: {e}")
            return False

    def send_key_combo(self, key1, key2):
        """단축키 전송 (예: Ctrl + F, Ctrl + V)"""
        pyautogui.hotkey(key1, key2)
        time.sleep(0.2)

    def press_key(self, key_name, count=1, delay=0.2):
        """단일 키 입력 (예: enter, esc, down 등)"""
        for _ in range(count):
            pyautogui.press(key_name)
            time.sleep(delay)

    def search_and_open_chat(self, target_name):
        """
        카카오톡에서 친구/채팅방을 검색하여 대화창을 엽니다.
        1. 카카오톡 메인 창 활성화
        2. Ctrl + F 로 검색창 진입
        3. 이름/전화번호 복사 후 붙여넣기
        4. Enter 키로 1번째 검색 결과 채팅방 열기
        """
        main_hwnd = self.find_kakao_main_window()
        if not main_hwnd:
            self.log("❌ 카카오톡이 실행되어 있지 않습니다. 카카오톡을 켜고 로그인해 주세요.")
            return False

        # 메인창 포커스
        self.bring_to_front(main_hwnd)
        time.sleep(0.4)

        # 친구 목록 탭으로 확실히 이동하기 위해 메인창에 ESC 두 번 후 Ctrl+F (검색창 열기)
        self.press_key('esc', count=2, delay=0.15)
        self.send_key_combo('ctrl', 'f')
        time.sleep(0.3)

        # 검색창 내용 지우기 (Ctrl+A -> Backspace)
        self.send_key_combo('ctrl', 'a')
        self.press_key('backspace')
        time.sleep(0.2)

        # 검색어 클립보드 복사 후 붙여넣기
        pyperclip.copy(target_name.strip())
        time.sleep(0.15)
        self.send_key_combo('ctrl', 'v')
        time.sleep(0.6)  # 검색 결과가 뜰 때까지 안전 대기

        # 1번째 검색 결과 선택 및 채팅방 열기 (Enter)
        self.press_key('enter')
        time.sleep(0.8)  # 채팅창이 뜰 때까지 대기

        return True

    def send_message_to_active_chat(self, message_text):
        """
        현재 열려 있는 채팅창에 클립보드를 이용해 메시지를 붙여넣고 전송합니다.
        1. 메시지 클립보드 복사
        2. Ctrl + V 붙여넣기
        3. Enter 키로 전송
        4. ESC 키로 대화창 닫기
        """
        # 메시지 클립보드 복사
        pyperclip.copy(message_text)
        time.sleep(0.2)

        # 대화창에 붙여넣기
        self.send_key_combo('ctrl', 'v')
        time.sleep(0.3)

        # 메시지 전송 (Enter)
        self.press_key('enter')
        time.sleep(0.4)

        # 채팅창 닫기 (ESC)
        self.press_key('esc')
        time.sleep(0.3)
        return True


class KakaoManagerGUI:
    """카카오톡 고객 관리 프로그램의 메인 대시보드 UI"""

    def __init__(self, root):
        self.root = root
        self.root.title("카카오톡 고객 안전 자동관리 프로그램 v2.0")
        self.root.geometry("860x700")
        self.root.minsize(800, 620)
        self.root.configure(bg="#1E1E2E")

        # 엔진 및 상태 변수 초기화
        self.engine = KakaoAutomationEngine(log_callback=self.append_log)
        self.worker_thread = None
        self.config_data = self.load_config()

        # UI 스타일 설정
        self.setup_styles()

        # UI 위젯 빌드
        self.create_widgets()

        # 엑셀 파일 로드 및 카운트 갱신
        self.refresh_excel_stats()

    def setup_styles(self):
        """모던 다크 테마 스타일 정의"""
        self.style = ttk.Style()
        self.style.theme_use("clam")

        # 기본 라벨 및 프레임
        self.style.configure("Dark.TFrame", background="#1E1E2E")
        self.style.configure("Card.TFrame", background="#282A36", relief="flat")
        self.style.configure("Dark.TLabel", background="#1E1E2E", foreground="#F8F8F2", font=("Malgun Gothic", 10))
        self.style.configure("Title.TLabel", background="#1E1E2E", foreground="#50FA7B", font=("Malgun Gothic", 15, "bold"))
        self.style.configure("SubTitle.TLabel", background="#1E1E2E", foreground="#8BE9FD", font=("Malgun Gothic", 9))

        # 카운터 카드 스타일
        self.style.configure("StatNum.TLabel", background="#282A36", foreground="#F8F8F2", font=("Malgun Gothic", 20, "bold"))
        self.style.configure("StatTitle.TLabel", background="#282A36", foreground="#6272A4", font=("Malgun Gothic", 9))

    def load_config(self):
        """config.json 설정 파일 불러오기"""
        default_config = {
            "app_password": "",
            "send_mode": "chat_folder",
            "direct_msg": "안녕하세요 {고객명} 고객님, 신한라이프 강동우입니다.\n좋은 하루 되세요!",
            "excel_path": "고객명단_양식.xlsx",
            "delay_min": 2.0,
            "delay_max": 4.0
        }
        if os.path.exists(CONFIG_FILE):
            try:
                with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    default_config.update(data)
            except Exception as e:
                print(f"설정 파일 읽기 오류: {e}")
        return default_config

    def save_config(self):
        """UI 입력값을 config.json에 저장"""
        try:
            self.config_data["direct_msg"] = self.txt_default_msg.get("1.0", tk.END).strip()
            self.config_data["excel_path"] = self.entry_excel_path.get().strip()
            self.config_data["delay_min"] = float(self.entry_delay_min.get().strip())
            self.config_data["delay_max"] = float(self.entry_delay_max.get().strip())

            with open(CONFIG_FILE, "w", encoding="utf-8") as f:
                json.dump(self.config_data, f, ensure_ascii=False, indent=4)
            self.append_log("💾 설정이 config.json에 안전하게 저장되었습니다.")
            messagebox.showinfo("저장 완료", "설정이 성공적으로 저장되었습니다!")
        except Exception as e:
            messagebox.showerror("저장 오류", f"설정 저장 중 오류가 발생했습니다:\n{e}")

    def create_widgets(self):
        """메인 레이아웃 위젯 생성"""
        main_frame = ttk.Frame(self.root, style="Dark.TFrame", padding=15)
        main_frame.pack(fill=tk.BOTH, expand=True)

        # 1. 헤더 영역
        header_frame = ttk.Frame(main_frame, style="Dark.TFrame")
        header_frame.pack(fill=tk.X, pady=(0, 10))

        lbl_title = ttk.Label(header_frame, text="💬 카카오톡 고객 안전 자동관리 프로그램 v2.0", style="Title.TLabel")
        lbl_title.pack(anchor="w")

        lbl_sub = ttk.Label(header_frame, text="노트북/데스크톱 화면 배율 완벽 대응 | 엑셀 맞춤 문구 치환 | 안전 지연 전송", style="SubTitle.TLabel")
        lbl_sub.pack(anchor="w", pady=(2, 0))

        # 2. 통계 카드 프레임 (4개 카드: 전체 대상, 발송 성공, 발송 실패, 대기)
        stat_frame = ttk.Frame(main_frame, style="Dark.TFrame")
        stat_frame.pack(fill=tk.X, pady=(0, 12))

        self.lbl_stat_total = self.create_stat_card(stat_frame, "📊 총 발송 대상", "0명", "#BD93F9", 0)
        self.lbl_stat_success = self.create_stat_card(stat_frame, "✅ 발송 성공", "0건", "#50FA7B", 1)
        self.lbl_stat_fail = self.create_stat_card(stat_frame, "❌ 발송 실패", "0건", "#FF5555", 2)
        self.lbl_stat_wait = self.create_stat_card(stat_frame, "⏳ 남은 대기", "0건", "#8BE9FD", 3)

        # 3. 설정 패널 (엑셀 경로 및 발송 설정)
        config_card = tk.LabelFrame(main_frame, text=" [ 발송 및 엑셀 설정 ] ", bg="#282A36", fg="#F8F8F2", font=("Malgun Gothic", 9, "bold"), bd=1, relief="solid")
        config_card.pack(fill=tk.X, pady=(0, 10), ipady=5)

        # 엑셀 경로 선택
        row1 = tk.Frame(config_card, bg="#282A36")
        row1.pack(fill=tk.X, padx=10, pady=4)

        tk.Label(row1, text="엑셀 파일 경로:", bg="#282A36", fg="#F8F8F2", font=("Malgun Gothic", 9)).pack(side=tk.LEFT)
        self.entry_excel_path = tk.Entry(row1, bg="#1E1E2E", fg="#50FA7B", insertbackground="white", font=("Malgun Gothic", 9))
        self.entry_excel_path.insert(0, self.config_data.get("excel_path", "고객명단_양식.xlsx"))
        self.entry_excel_path.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=8)

        btn_find_excel = tk.Button(row1, text="📁 찾아보기", bg="#44475A", fg="#F8F8F2", font=("Malgun Gothic", 8), command=self.browse_excel, relief="flat")
        btn_find_excel.pack(side=tk.LEFT, padx=2)

        btn_open_excel = tk.Button(row1, text="📊 엑셀 열기", bg="#44475A", fg="#F8F8F2", font=("Malgun Gothic", 8), command=self.open_excel_file, relief="flat")
        btn_open_excel.pack(side=tk.LEFT, padx=2)

        # 딜레이 및 설정
        row2 = tk.Frame(config_card, bg="#282A36")
        row2.pack(fill=tk.X, padx=10, pady=4)

        tk.Label(row2, text="안전 딜레이(초):", bg="#282A36", fg="#F8F8F2", font=("Malgun Gothic", 9)).pack(side=tk.LEFT)
        self.entry_delay_min = tk.Entry(row2, width=5, bg="#1E1E2E", fg="#F8F8F2", justify="center")
        self.entry_delay_min.insert(0, str(self.config_data.get("delay_min", 2.0)))
        self.entry_delay_min.pack(side=tk.LEFT, padx=4)

        tk.Label(row2, text="~", bg="#282A36", fg="#F8F8F2").pack(side=tk.LEFT)
        self.entry_delay_max = tk.Entry(row2, width=5, bg="#1E1E2E", fg="#F8F8F2", justify="center")
        self.entry_delay_max.insert(0, str(self.config_data.get("delay_max", 4.0)))
        self.entry_delay_max.pack(side=tk.LEFT, padx=4)

        tk.Label(row2, text="초 (계정보호를 위해 랜덤 지연 발송)", bg="#282A36", fg="#6272A4", font=("Malgun Gothic", 8)).pack(side=tk.LEFT, padx=6)

        btn_save_config = tk.Button(row2, text="💾 설정 저장", bg="#6272A4", fg="#F8F8F2", font=("Malgun Gothic", 8, "bold"), command=self.save_config, relief="flat")
        btn_save_config.pack(side=tk.RIGHT)

        # 기본 발송 문구 (템플릿)
        row3 = tk.Frame(config_card, bg="#282A36")
        row3.pack(fill=tk.X, padx=10, pady=4)
        tk.Label(row3, text="기본 발송 문구 (엑셀의 '전송할메시지'가 비어있을 때 사용되며, {고객명}은 자동 치환됩니다):", bg="#282A36", fg="#8BE9FD", font=("Malgun Gothic", 8)).pack(anchor="w")

        self.txt_default_msg = tk.Text(row3, height=4, bg="#1E1E2E", fg="#F8F8F2", insertbackground="white", font=("Malgun Gothic", 9), wrap=tk.WORD)
        self.txt_default_msg.insert("1.0", self.config_data.get("direct_msg", ""))
        self.txt_default_msg.pack(fill=tk.X, pady=(2, 0))

        # 4. 제어 버튼 바 (시작 / 일시정지 / 중지)
        btn_bar = tk.Frame(main_frame, bg="#1E1E2E")
        btn_bar.pack(fill=tk.X, pady=(0, 10))

        self.btn_start = tk.Button(btn_bar, text="🚀 자동 발송 시작", bg="#50FA7B", fg="#1E1E2E", font=("Malgun Gothic", 11, "bold"), height=2, command=self.start_automation, relief="flat")
        self.btn_start.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 4))

        self.btn_pause = tk.Button(btn_bar, text="⏸️ 일시 정지", bg="#FFB86C", fg="#1E1E2E", font=("Malgun Gothic", 11, "bold"), height=2, command=self.toggle_pause, relief="flat", state=tk.DISABLED)
        self.btn_pause.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=4)

        self.btn_stop = tk.Button(btn_bar, text="⏹️ 작업 중지", bg="#FF5555", fg="#FFFFFF", font=("Malgun Gothic", 11, "bold"), height=2, command=self.stop_automation, relief="flat", state=tk.DISABLED)
        self.btn_stop.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(4, 0))

        # 5. 실시간 진행 상황 및 로그 영역
        log_frame = tk.LabelFrame(main_frame, text=" [ 실시간 진행 상황 및 로그 ] ", bg="#282A36", fg="#F8F8F2", font=("Malgun Gothic", 9, "bold"), bd=1, relief="solid")
        log_frame.pack(fill=tk.BOTH, expand=True)

        self.txt_log = tk.Text(log_frame, bg="#11111B", fg="#50FA7B", insertbackground="white", font=("Consolas", 9), wrap=tk.WORD)
        self.txt_log.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)

        scrollbar = ttk.Scrollbar(log_frame, orient=tk.VERTICAL, command=self.txt_log.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.txt_log.config(yscrollcommand=scrollbar.set)

        self.append_log("✨ 프로그램이 정상적으로 로드되었습니다.")
        self.append_log("📌 카카오톡에 로그인 후 [🚀 자동 발송 시작] 버튼을 눌러주세요.")

    def create_stat_card(self, parent, title, initial_val, color, col_idx):
        """통계 카드 생성 유틸"""
        card = tk.Frame(parent, bg="#282A36", bd=1, relief="solid", padx=10, pady=8)
        card.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=3)

        lbl_t = tk.Label(card, text=title, bg="#282A36", fg="#6272A4", font=("Malgun Gothic", 9))
        lbl_t.pack(anchor="w")

        lbl_v = tk.Label(card, text=initial_val, bg="#282A36", fg=color, font=("Malgun Gothic", 15, "bold"))
        lbl_v.pack(anchor="w", pady=(3, 0))
        return lbl_v

    def append_log(self, text):
        """로그 텍스트 추가 (스레드 안전)"""
        def _append():
            self.txt_log.insert(tk.END, text + "\n")
            self.txt_log.see(tk.END)
        self.root.after(0, _append)

    def browse_excel(self):
        """엑셀 파일 찾아보기 다이얼로그"""
        filename = filedialog.askopenfilename(
            title="고객 명단 엑셀 파일 선택",
            filetypes=[("Excel Files", "*.xlsx *.xls"), ("All Files", "*.*")],
            initialdir=BASE_DIR
        )
        if filename:
            self.entry_excel_path.delete(0, tk.END)
            self.entry_excel_path.insert(0, filename)
            self.refresh_excel_stats()

    def open_excel_file(self):
        """엑셀 파일 바로 열기"""
        path = self.entry_excel_path.get().strip()
        if not os.path.isabs(path):
            path = os.path.join(BASE_DIR, path)
        if os.path.exists(path):
            os.startfile(path)
        else:
            messagebox.showerror("오류", f"엑셀 파일을 찾을 수 없습니다:\n{path}")

    def refresh_excel_stats(self):
        """엑셀 파일을 읽어 전체 대상 및 발송 완료 건수 파악"""
        path = self.entry_excel_path.get().strip()
        if not os.path.isabs(path):
            path = os.path.join(BASE_DIR, path)

        if not os.path.exists(path):
            self.lbl_stat_total.config(text="0명")
            self.lbl_stat_success.config(text="0건")
            self.lbl_stat_fail.config(text="0건")
            self.lbl_stat_wait.config(text="0건")
            return

        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            sheet = wb.active
            rows = list(sheet.iter_rows(values_only=True))
            if len(rows) <= 1:
                return

            total_count = 0
            success_count = 0
            fail_count = 0
            wait_count = 0

            for row in rows[1:]:
                # 이름이나 연락처가 있는 행만 카운트
                if row and (row[0] or (len(row) > 1 and row[1])):
                    total_count += 1
                    status = str(row[3]).strip() if len(row) > 3 and row[3] else ""
                    if status == "성공":
                        success_count += 1
                    elif status == "실패":
                        fail_count += 1
                    else:
                        wait_count += 1

            self.lbl_stat_total.config(text=f"{total_count}명")
            self.lbl_stat_success.config(text=f"{success_count}건")
            self.lbl_stat_fail.config(text=f"{fail_count}건")
            self.lbl_stat_wait.config(text=f"{wait_count}건")
        except Exception as e:
            self.append_log(f"⚠️ 엑셀 상태 갱신 알림: {e}")

    def start_automation(self):
        """자동 발송 시작"""
        # 카카오톡 실행 여부 확인
        hwnd = self.engine.find_kakao_main_window()
        if not hwnd:
            messagebox.showwarning("카카오톡 미실행", "카카오톡 PC버전이 실행되어 있지 않습니다.\n카카오톡을 켜고 로그인하신 후 다시 시작해 주세요.")
            return

        self.engine.is_running = True
        self.engine.is_paused = False

        self.btn_start.config(state=tk.DISABLED, bg="#44475A")
        self.btn_pause.config(state=tk.NORMAL, text="⏸️ 일시 정지")
        self.btn_stop.config(state=tk.NORMAL)

        # 백그라운드 스레드로 발송 작업 실행
        self.worker_thread = threading.Thread(target=self.run_process, daemon=True)
        self.worker_thread.start()

    def toggle_pause(self):
        """일시정지 / 재개 토글"""
        if self.engine.is_paused:
            self.engine.is_paused = False
            self.btn_pause.config(text="⏸️ 일시 정지")
            self.append_log("▶️ 작업이 다시 재개되었습니다.")
        else:
            self.engine.is_paused = True
            self.btn_pause.config(text="▶️ 작업 재개")
            self.append_log("⏸️ 작업이 일시 정지되었습니다.")

    def stop_automation(self):
        """작업 중지"""
        self.engine.is_running = False
        self.engine.is_paused = False
        self.append_log("⏹️ 작업 중지를 요청했습니다...")

    def run_process(self):
        """실제 발송 루프 (백그라운드 스레드)"""
        excel_rel_path = self.entry_excel_path.get().strip()
        excel_full_path = excel_rel_path if os.path.isabs(excel_rel_path) else os.path.join(BASE_DIR, excel_rel_path)

        if not os.path.exists(excel_full_path):
            self.append_log(f"❌ 엑셀 파일을 찾을 수 없습니다: {excel_full_path}")
            self.reset_ui_buttons()
            return

        delay_min = float(self.entry_delay_min.get().strip())
        delay_max = float(self.entry_delay_max.get().strip())
        default_msg_template = self.txt_default_msg.get("1.0", tk.END).strip()

        self.append_log("==================================================")
        self.append_log("🚀 카카오톡 고객 자동 발송을 시작합니다.")
        self.append_log(f"📄 대상 엑셀: {os.path.basename(excel_full_path)}")
        self.append_log(f"⏱️ 지연 시간: {delay_min}초 ~ {delay_max}초")
        self.append_log("==================================================")

        try:
            wb = openpyxl.load_workbook(excel_full_path)
            sheet = wb.active
            rows = list(sheet.iter_rows())

            if len(rows) <= 1:
                self.append_log("⚠️ 엑셀에 발송할 고객 데이터가 없습니다.")
                self.reset_ui_buttons()
                return

            total_targets = len(rows) - 1
            current_idx = 0

            for row_idx in range(2, len(rows) + 1):
                if not self.engine.is_running:
                    self.append_log("⏹️ 사용자에 의해 작업이 안전하게 중단되었습니다.")
                    break

                # 일시 정지 상태 처리
                while self.engine.is_paused:
                    time.sleep(0.5)
                    if not self.engine.is_running:
                        break

                current_idx += 1
                cust_name = str(sheet.cell(row=row_idx, column=1).value or "").strip()
                kakao_target = str(sheet.cell(row=row_idx, column=2).value or "").strip()
                custom_msg = str(sheet.cell(row=row_idx, column=3).value or "").strip()
                status = str(sheet.cell(row=row_idx, column=4).value or "").strip()

                # 빈 행이면 건너뛰기
                if not cust_name and not kakao_target:
                    continue

                # 이미 성공한 건은 중복 발송 방지
                if status == "성공":
                    self.append_log(f"[{current_idx}/{total_targets}] '{cust_name}'님은 이미 전송 완료되어 건너뜁니다.")
                    continue

                # 검색 대상 결정 (카톡이름/전화번호 우선, 없으면 고객성함)
                search_query = kakao_target if kakao_target else cust_name

                # 최종 발송 메시지 구성 (개인화 치환)
                if custom_msg:
                    final_msg = custom_msg
                else:
                    final_msg = default_msg_template.replace("{고객명}", cust_name).replace("{이름}", cust_name)

                self.append_log(f"[{current_idx}/{total_targets}] '{cust_name}' ({search_query}) 검색 및 전송 시도 중...")

                # 1. 카카오톡 검색 및 채팅방 열기
                open_success = self.engine.search_and_open_chat(search_query)

                now_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                if open_success:
                    # 2. 메시지 전송
                    send_success = self.engine.send_message_to_active_chat(final_msg)
                    if send_success:
                        sheet.cell(row=row_idx, column=4, value="성공")
                        sheet.cell(row=row_idx, column=5, value=now_time)
                        sheet.cell(row=row_idx, column=6, value="전송 완료")
                        self.append_log(f"  └ ✅ 전송 완료!")
                    else:
                        sheet.cell(row=row_idx, column=4, value="실패")
                        sheet.cell(row=row_idx, column=5, value=now_time)
                        sheet.cell(row=row_idx, column=6, value="메시지 입력 실패")
                        self.append_log(f"  └ ❌ 메시지 입력 실패")
                else:
                    sheet.cell(row=row_idx, column=4, value="실패")
                    sheet.cell(row=row_idx, column=5, value=now_time)
                    sheet.cell(row=row_idx, column=6, value="친구/채팅방 검색 실패")
                    self.append_log(f"  └ ❌ 검색 결과 없음 / 채팅방 열기 실패")

                # 실시간 엑셀 파일 저장
                try:
                    wb.save(excel_full_path)
                except Exception as save_err:
                    self.append_log(f"  └ ⚠️ 엑셀 저장 알림 (파일이 열려있으면 닫아주세요): {save_err}")

                # 통계 UI 업데이트
                self.root.after(0, self.refresh_excel_stats)

                # 안전 지연 대기 (마지막 행이 아니면)
                if row_idx < len(rows) and self.engine.is_running:
                    sleep_time = random.uniform(delay_min, delay_max)
                    self.append_log(f"  └ ⏳ 계정 보호를 위해 {sleep_time:.1f}초 안전 대기 중...")
                    time.sleep(sleep_time)

            self.append_log("==================================================")
            self.append_log("🎉 모든 대상에 대한 발송 작업이 완료되었습니다!")
            self.append_log("==================================================")

        except Exception as e:
            self.append_log(f"❌ 작업 도중 예외 발생: {e}")
        finally:
            self.reset_ui_buttons()
            self.root.after(0, self.refresh_excel_stats)

    def reset_ui_buttons(self):
        """버튼 상태 원복 (스레드 안전)"""
        def _reset():
            self.engine.is_running = False
            self.engine.is_paused = False
            self.btn_start.config(state=tk.NORMAL, bg="#50FA7B")
            self.btn_pause.config(state=tk.DISABLED, text="⏸️ 일시 정지")
            self.btn_stop.config(state=tk.DISABLED)
        self.root.after(0, _reset)


if __name__ == "__main__":
    root = tk.Tk()
    app = KakaoManagerGUI(root)
    root.mainloop()